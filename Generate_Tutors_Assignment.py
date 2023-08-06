import openpyxl
import random
import pandas as pd

excel_path_processing = ('cle/TutoringTranscriptSampleProcessed.xlsx')
excel_path_final_result = ('cle/Tutors Assignment.xlsx')

maximum_rows = 50

# give each tutor a random selection of courses from the list of courses
def randomize(tutor):
    random_courses = random.sample(courses, 5)
    random_courses.insert(0,tutor)
    biglist.append(random_courses)


# increase the selected slot in excel from one to the next, ex: from A1 to A2 and so on
def through(i,j,k):
    list1 = list()
    startC = chr(k)
    for _ in range(i,j):
        startN = 0
        for _ in range(maximum_rows):
            startN += 1
            cursor = startC+str(startN)
            list1.append(cursor)
        startC = chr(ord(startC) + 1)
    return list1

# fill the fields in the excel file with the selected courses and tutors using the previous function: i and j: how many columns to fill, k: which column to start filling from, l: which information to feed into the excel file (tutor name, course 1, course 2, etc)
def gothrough(sheet, i, j, k, l):
    orderlist = through(i, j, k)
    innerlist = [sub_list[l] for sub_list in biglist]
    for element1, element2 in zip(orderlist, innerlist):
        sheet[element1] = element2

# list of what would be inserted in the excel file and the titles of the columns
biglist = list()
biglist.append(['Tutor', 'Course1', 'Course2', 'Course3', 'Course4', 'Course5'])

# create lists to be used later
list_of_everything = list()
courses = list()
tutors = list()

# open the excel file and select the sheets to work with/fill 
wb = openpyxl.load_workbook(excel_path_processing)
wb_final = openpyxl.Workbook()
sheet_final = wb_final.active

# Select the desired column to which we would like to know the maximum number of rows
column_name = 'Course_Title'  

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Read the Excel file into a DataFrame
    df = pd.read_excel(excel_path_processing, engine='openpyxl', sheet_name=sheet_name)

    # Get the maximum number of rows in the column (column_name) specified before
    max_rows = df[column_name].last_valid_index() + 1


    # get the list of tutors and courses in one big list
    for value in ws.iter_rows(
        min_row=2, max_row=max_rows+1, min_col=1, max_col=2, values_only=True):
        list_of_everything.append(value)

    # divide the big list into courses and tutors 
    for tutor, course in list_of_everything:
        courses.append(course)
        tutors.append(tutor)
    
    # assign for each tutor 5 random courses
    randomize(tutor)

    # insert the results in the excel file, i and j: how many columns to fill, k: which column to start filling from, l: which information to feed into the excel file (tutor name, course 1, course 2, etc)
    for i, j, k, l in zip(range(2,8), range(3,9), range(67,73), range(6)):
        gothrough(ws, i, j, k, l)
        gothrough(sheet_final, i, j, k-2, l)
    
    list_of_everything = list()
    courses.clear()

# save the changes to the excel files
wb.save(excel_path_processing)
wb_final.save(excel_path_final_result)